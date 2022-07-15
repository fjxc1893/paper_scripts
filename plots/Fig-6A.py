# -*- coding: utf-8 -*-
# Time : 2021/12/6 8:42
# Author : 张志康

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib
import openpyxl


##########################################
workbook = openpyxl.load_workbook("Mesenchymal vs others可信蛋白 protein-final-add-GO+KEGG.xlsx")
worksheet = workbook["sheet 2-细胞外基质卡方检验（M918T和蛋白表达）"]

# 获取当前表单中单元格的行列数
rows, cols = worksheet.max_row, worksheet.max_column
genes = []
#找到第三列是红色字体的行
for i in range(1,rows+1):
    if worksheet.cell(i, 3).font.color.rgb == "FFFF0000":
        genes.append(worksheet.cell(i, 3).value)
####################################################
def fdr(p_vals):
    from scipy.stats import rankdata
    ranked_p_values = rankdata(p_vals)
    fdr = p_vals * len(p_vals) / ranked_p_values
    fdr[fdr > 1] = 1
    return fdr


df = pd.read_excel("Mesenchymal vs others可信蛋白 protein-final-add-GO+KEGG.xlsx",sheet_name="Mesenchymal vs others 差异-添加注释信息")
df = df[["Accession","Gene Name","P-value","log2(FC)","ECM","Secreted"]]
df["Q-value"] = fdr(df["P-value"])
df["-log10(qvalue)"] = df["Q-value"].map(lambda x:-np.log10(x))
print(df)


Collagens_Secreted = df[(df["ECM"] == "Collagens") & (df["Secreted"] == "Secreted")]
Collagens_NonSecreted = df[(df["ECM"] == "Collagens") & (df["Secreted"] != "Secreted")]

ECM_affiliated_Proteins_Secreted = df[(df["ECM"] == "ECM-affiliated Proteins") & (df["Secreted"] == "Secreted")]
ECM_affiliated_Proteins_NonSecreted = df[(df["ECM"] == "ECM-affiliated Proteins") & (df["Secreted"] != "Secreted")]

ECM_Glycoproteins_Secreted = df[(df["ECM"] == "ECM Glycoproteins") & (df["Secreted"] == "Secreted") ]
ECM_Glycoproteins_NonSecreted = df[(df["ECM"] == "ECM Glycoproteins") & (df["Secreted"] != "Secreted") ]

ECM_Regulators_Secreted = df[(df["ECM"] == "ECM Regulators") & (df["Secreted"] == "Secreted")]
ECM_Regulators_NonSecreted = df[(df["ECM"] == "ECM Regulators") & (df["Secreted"] != "Secreted")]

Proteoglycans_Secreted = df[(df["ECM"] == "Proteoglycans") & (df["Secreted"] == "Secreted")]
Proteoglycans_NonSecreted = df[(df["ECM"] == "Proteoglycans") & (df["Secreted"] != "Secreted")]

Secreted_Factors_Secreted = df[(df["ECM"] == "Secreted Factors") & (df["Secreted"] == "Secreted")]
Secreted_Factors_NonSecreted = df[(df["ECM"] == "Secreted Factors") & (df["Secreted"] != "Secreted")]

TNC = df[df["Gene Name"] == 'TNC'][["log2(FC)","-log10(qvalue)"]].values[0]
TNXB = df[df["Gene Name"] == 'TNXB'][["log2(FC)","-log10(qvalue)"]].values[0]
#####################################
plt.rcParams['font.sans-serif'] = ['Arial']
plt.rcParams['axes.unicode_minus'] = False
######################################
fig = plt.figure(figsize=(8, 8))
ax = fig.add_axes([0.1, 0.1, 0.8, 0.8])

line = plt.axhline(y=-np.log10(0.05),ls="--",color="grey")
# line.set_zorder(-100)

ax.scatter(df["log2(FC)"],df["-log10(qvalue)"],color="#BFBFBF")

ax.scatter(Collagens_Secreted["log2(FC)"],Collagens_Secreted["-log10(qvalue)"],color="#ABCD9A",alpha=0.9,label="Collagens",marker="^")
ax.scatter(Collagens_NonSecreted["log2(FC)"],Collagens_NonSecreted["-log10(qvalue)"],color="#ABCD9A",alpha=0.9,label="Collagens")
ax.scatter(ECM_affiliated_Proteins_Secreted["log2(FC)"],ECM_affiliated_Proteins_Secreted["-log10(qvalue)"],color="#F2C6B3",alpha=0.9,marker="^")
ax.scatter(ECM_affiliated_Proteins_NonSecreted["log2(FC)"],ECM_affiliated_Proteins_NonSecreted["-log10(qvalue)"],color="#F2C6B3",alpha=0.9)
ax.scatter(ECM_Glycoproteins_Secreted["log2(FC)"],ECM_Glycoproteins_Secreted["-log10(qvalue)"],color="#7AACDA",alpha=0.9,marker="^")
ax.scatter(ECM_Glycoproteins_NonSecreted["log2(FC)"],ECM_Glycoproteins_NonSecreted["-log10(qvalue)"],color="#7AACDA",alpha=0.9)
ax.scatter(ECM_Regulators_Secreted["log2(FC)"],ECM_Regulators_Secreted["-log10(qvalue)"],color="#FFE28A",alpha=0.9,marker="^")
ax.scatter(ECM_Regulators_NonSecreted["log2(FC)"],ECM_Regulators_NonSecreted["-log10(qvalue)"],color="#FFE28A",alpha=0.9)
ax.scatter(Proteoglycans_Secreted["log2(FC)"],Proteoglycans_Secreted["-log10(qvalue)"],color="#FAD6FA",alpha=0.9,marker="^")
ax.scatter(Proteoglycans_NonSecreted["log2(FC)"],Proteoglycans_NonSecreted["-log10(qvalue)"],color="#FAD6FA",alpha=0.9)
ax.scatter(Secreted_Factors_Secreted["log2(FC)"],Secreted_Factors_Secreted["-log10(qvalue)"],color="#ADAD5A",alpha=0.9,marker="^")
ax.scatter(Secreted_Factors_NonSecreted["log2(FC)"],Secreted_Factors_NonSecreted["-log10(qvalue)"],color="#ADAD5A",alpha=0.9)

#############################################################################################################
plt.annotate("TNC",xy=TNC,xytext=(TNC[0]-0.08,TNC[1]-0.7),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},color="red",weight="bold",size=12)
plt.annotate("TNXB",xy=TNXB,xytext=(TNXB[0]+0.17,TNXB[1]-0.12),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},color="red",weight="bold",size=12)



genes_ = [
    'ANXA7',
    'ANXA9',
    'ANXA5',
    'CSTB',
    'CST3',
    'CTF1',
    'S100A1',
    'S100A7',
    'SPARCL1',
    'CTSO',
    'FBLN2',
    'THBS3',
    'A2M',
    'ITIH4',
    'FCN3',
    'EFEMP2',
    'LAMA4',
    'COL16A1',
    'COL5A2',
    'SFTPB',
    'NPNT',
    'NID2',
    'ADIPOQ',
    'HABP2',
    'CTSS',
    'CD109',
    'PLG',
    'MFAP4',
    'PLXDC2',
    'LAMA2'
]

# print(set(genes) & set(genes_))

df_ = df[(df["Gene Name"].isin(genes)) & (df["log2(FC)"] > 0)]
# print(df_[["Gene Name","log2(FC)","-log10(qvalue)"]].sort_values(by="-log10(qvalue)"))

#print(ECM_Regulators_Secreted[(ECM_Regulators_Secreted["-log10(qvalue)"] > 1) & (ECM_Regulators_Secreted["-log10(qvalue)"] < 2)])

ANXA7 = df[df["Gene Name"] == 'ANXA7'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("ANXA7",xy=ANXA7,xytext=(ANXA7[0]+0.3,ANXA7[1]+0.3),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=50,armB=0,rad=0"},size=10)

ANXA9 = df[df["Gene Name"] == 'ANXA9'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("ANXA9",xy=ANXA9,xytext=(ANXA9[0]-0.4,ANXA9[1]+0.5),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=-180,angleB=0,armA=50,armB=0,rad=0"},size=10)
plt.annotate("ANXA9",xy=ANXA9,xytext=(ANXA9[0]-0.4,ANXA9[1]-0.35),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-30,armB=0,rad=0"},size=10)

ANXA5 = df[df["Gene Name"] == 'ANXA5'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("ANXA5",xy=ANXA5,xytext=(ANXA5[0]-0.7,ANXA5[1]+2.5),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

CSTB = df[df["Gene Name"] == 'CSTB'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("CSTB",xy=CSTB,xytext=(CSTB[0]-0.6,CSTB[1]+0.8),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

CST3 = df[df["Gene Name"] == 'CST3'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("CST3",xy=CST3,xytext=(CST3[0]-0.6,CST3[1]+0.8),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

CTF1 = df[df["Gene Name"] == 'CTF1'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("CTF1",xy=CTF1,xytext=(CTF1[0]-0.55,CTF1[1]-0.2),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

S100A1 = df[df["Gene Name"] == 'S100A1'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("S100A1",xy=S100A1,xytext=(S100A1[0]-0.55,S100A1[1]-0.45),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

S100A7 = df[df["Gene Name"] == 'S100A7'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("S100A7",xy=S100A7,xytext=(S100A7[0]-0.2,S100A7[1]-1.05),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

SPARCL1 = df[df["Gene Name"] == 'SPARCL1'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("SPARCL1",xy=SPARCL1,xytext=(SPARCL1[0]-0.7,SPARCL1[1]-0.65),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-45,armB=0,rad=0"},size=10)

################################################################
# CTSO = mRNA[mRNA["Gene Name"] == 'CTSO'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("CTSO",xy=CTSO,xytext=(CTSO[0]-0.6,CTSO[1]+1),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-50,armB=0,rad=0"},size=10)

# FBLN2 = mRNA[mRNA["Gene Name"] == 'FBLN2'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("FBLN2",xy=FBLN2,xytext=(FBLN2[0]+0.16,FBLN2[1]+1),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

THBS3 = df[df["Gene Name"] == 'THBS3'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("THBS3",xy=THBS3,xytext=(THBS3[0]-0.7,THBS3[1]+1.95),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

A2M = df[df["Gene Name"] == 'A2M'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("A2M",xy=A2M,xytext=(A2M[0]+0.3,A2M[1]+1.2),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

ITIH4 = df[df["Gene Name"] == 'ITIH4'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("ITIH4",xy=ITIH4,xytext=(ITIH4[0]-0.7,ITIH4[1]+1.5),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-30,armB=0,rad=0"},size=10)

FCN3 = df[df["Gene Name"] == 'FCN3'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("FCN3",xy=FCN3,xytext=(FCN3[0]-0.8,FCN3[1]+1.3),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-35,armB=0,rad=0"},size=10)

EFEMP2 = df[df["Gene Name"] == 'EFEMP2'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("EFEMP2",xy=EFEMP2,xytext=(EFEMP2[0]-0.8,EFEMP2[1]+1.1),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

LAMA4 = df[df["Gene Name"] == 'LAMA4'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("LAMA4",xy=LAMA4,xytext=(LAMA4[0]-0.6,LAMA4[1]+0.8),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

COL16A1 = df[df["Gene Name"] == 'COL16A1'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("COL16A1",xy=COL16A1,xytext=(COL16A1[0]-0.8,COL16A1[1]+1.1),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

# COL6A1 = mRNA[mRNA["Gene Name"] == 'COL6A1'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("COL6A1",xy=COL6A1,xytext=(COL6A1[0]+0.5,COL6A1[1]-2),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

# COL5A2 = mRNA[mRNA["Gene Name"] == 'COL5A2'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("COL5A2",xy=COL5A2,xytext=(COL5A2[0]+0.1,COL5A2[1]-0.4),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

# SFTPB = mRNA[mRNA["Gene Name"] == 'SFTPB'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("SFTPB",xy=SFTPB,xytext=(SFTPB[0]+0.15,SFTPB[1]-0.4),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)


# NPNT = mRNA[mRNA["Gene Name"] == 'NPNT'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("NPNT",xy=NPNT,xytext=(NPNT[0]-0.4,NPNT[1]+0.4),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

# NID2 = mRNA[mRNA["Gene Name"] == 'NID2'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("NID2",xy=NID2,xytext=(NID2[0]-0.45,NID2[1]+0.45),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

# ADIPOQ = mRNA[mRNA["Gene Name"] == 'ADIPOQ'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("ADIPOQ",xy=ADIPOQ,xytext=(ADIPOQ[0]-0.6,ADIPOQ[1]+1.32),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-50,armB=0,rad=0"},size=10)

HABP2 = df[df["Gene Name"] == 'HABP2'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("HABP2",xy=HABP2,xytext=(HABP2[0]-0.5,HABP2[1]+0.5),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-30,armB=0,rad=0"},size=10)

CTSS = df[df["Gene Name"] == 'CTSS'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("CTSS",xy=CTSS,xytext=(CTSS[0]-0.55,CTSS[1]+0.9),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-30,armB=0,rad=0"},size=10)

# CD109 = mRNA[mRNA["Gene Name"] == 'CD109'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("CD109",xy=CD109,xytext=(CD109[0]+0.2,CD109[1]-1.1),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

PLG = df[df["Gene Name"] == 'PLG'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("PLG",xy=PLG,xytext=(PLG[0]-0.25,PLG[1]+0.25),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

# MFAP4 = mRNA[mRNA["Gene Name"] == 'MFAP4'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("MFAP4",xy=MFAP4,xytext=(MFAP4[0]+0.2,MFAP4[1]+0.8),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

# PLXDC2 = mRNA[mRNA["Gene Name"] == 'PLXDC2'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("PLXDC2",xy=PLXDC2,xytext=(PLXDC2[0]+0.2,PLXDC2[1]+0.5),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

# LAMA2 = mRNA[mRNA["Gene Name"] == 'LAMA2'][["log2(FC)","-log10(qvalue)"]].values[0]
# plt.annotate("LAMA2",xy=LAMA2,xytext=(LAMA2[0]+0.2,LAMA2[1]+0.2),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

#HGF = mRNA[mRNA["Gene Name"] == 'HGF'][["log2(FC)","-log10(qvalue)"]].values[0]
#plt.annotate("HGF",xy=HGF,xytext=(HGF[0]+0.08,HGF[1]-0.45),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

ECM1 = df[df["Gene Name"] == 'ECM1'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("ECM1",xy=ECM1,xytext=(ECM1[0]+0.2,ECM1[1]-0.4),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

ECM2 = df[df["Gene Name"] == 'ECM2'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("ECM2",xy=ECM2,xytext=(ECM2[0]+0.6,ECM2[1]-0.5),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=35,armB=0,rad=0"},size=10)

SLIT3 = df[df["Gene Name"] == 'SLIT3'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("SLIT3",xy=SLIT3,xytext=(SLIT3[0]-0.35,SLIT3[1]-0.1),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-20,armB=0,rad=0"},size=10)

COL4A2 = df[df["Gene Name"] == 'COL4A2'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("COL4A2",xy=COL4A2,xytext=(COL4A2[0]+0.7,COL4A2[1]-0.45),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

#F12 = mRNA[mRNA["Gene Name"] == 'F12'][["log2(FC)","-log10(qvalue)"]].values[0]
#plt.annotate("F12",xy=F12,xytext=(F12[0]-0.5,F12[1]+0.9),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

#FMOD = mRNA[mRNA["Gene Name"] == 'FMOD'][["log2(FC)","-log10(qvalue)"]].values[0]
#plt.annotate("FMOD",xy=FMOD,xytext=(FMOD[0]+0.15,FMOD[1]-0.8),bbox=dict(boxstyle="round,pad=0.5",fc='white'),arrowprops={'arrowstyle':'->'},size=10)

#FGB = mRNA[mRNA["Gene Name"] == 'FGB'][["log2(FC)","-log10(qvalue)"]].values[0]
#plt.annotate("FGB",xy=FGB,xytext=(FGB[0]+0.55,FGB[1]-0.79),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=30,armB=0,rad=0"},size=10)

COL3A1 = df[df["Gene Name"] == 'COL3A1'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("COL3A1",xy=COL3A1,xytext=(COL3A1[0]+0.12,COL3A1[1]+0.6),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->'},size=10)

VWF = df[df["Gene Name"] == 'VWF'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("VWF",xy=VWF,xytext=(VWF[0]+0.5,VWF[1]+0.65),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=40,armB=0,rad=0"},size=10)

SERPINA5 = df[df["Gene Name"] == 'SERPINA5'][["log2(FC)","-log10(qvalue)"]].values[0]
plt.annotate("SERPINA5",xy=SERPINA5,xytext=(SERPINA5[0]-0.55,SERPINA5[1]+0.23),bbox=dict(boxstyle="round,pad=0.5",fc='None'),arrowprops={'arrowstyle':'->',"connectionstyle":"arc,angleA=180,angleB=0,armA=-40,armB=0,rad=0"},size=10)

ax.set_xlabel(r"log$_{2}$FoldChange",size=20)
ax.set_ylabel(r"-log$_{10}$qvalue",size=20)
plt.xticks(size=15)
plt.yticks(size=15)
ax.spines['bottom'].set_linewidth(1.5)
ax.spines['left'].set_linewidth(1.5)
ax.spines['top'].set_linewidth(1.5)
ax.spines['right'].set_linewidth(1.5)
# plt.xlim(-1.7,1.7)

legends = [matplotlib.lines.Line2D([0],[0], linestyle="none", c=i, marker = '.',markersize=10) for i in ("#ABCD9A","#F2C6B3","#7AACDA","#FFE28A","#FAD6FA","#ADAD5A",)] + [matplotlib.lines.Line2D([0],[0], linestyle="none", c="black", marker = "^")]
labels = ["Collagens","ECM-affiliated Proteins","ECM Glycoproteins","ECM Regulators","Proteoglycans","Secreted Factors","Secreted"]

plt.legend(legends,labels,frameon=False)
######################
# with plt.rc_context({'image.composite_image': True}):
#     plt.savefig("volcano.pdf")
#     plt.savefig("volcano.png")
plt.savefig("volcano.pdf")
# plt.show()

